
import openpyxl,requests,json

dataframe = openpyxl.load_workbook("wt_number.xlsx")

activated_dataframe = dataframe.active


def add_to_whatsapp_group(number):
    
    grpupId = 'your whatsapp group id'
    url = f"https://gate.whapi.cloud/groups/{groupId}/participants"

    payload = { "participants": [number] }
    headers = {
        "accept": "application/json",
        "content-type": "application/json",
        "authorization": "Bearer tf3xU0wyrcQQ7UAIU5ZJti8VFySULiix"
    }

    response = requests.post(url, json=payload, headers=headers)
    print(response.status_code)
    # processed = json.loads(response.text).get('processed',[])
    # print(len(processed))
    print(response.text)
    return response

succeed=0
failed = 0

for row in range(1, activated_dataframe.max_row):
    for col in activated_dataframe.iter_cols(2, 2):
        # print(col[row].value)
        response = add_to_whatsapp_group(str(col[row].value))
        processed = json.loads(response.text).get('processed',[])
        if response.ok and len(processed)> 0:
            succeed+=1
            print(f"{col[row].value}: successfully added")
        elif response.status_code==409:
            print(f"{col[row].value}: already added")
            failed+=1
        else:
            print(f"{col[row].value}: failed for other reason {response.text}")
            failed+=1
    
print(f"total added: {succeed}, total failed: {failed}")
        

